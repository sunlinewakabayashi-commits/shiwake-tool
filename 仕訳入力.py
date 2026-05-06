"""
仕訳入力ツール
レシート画像を読み取り、あかしまい_損益確認 R8年度分.xlsx に自動入力します。

使い方: python 仕訳入力.py 会社名_月.jpg [ファイル2 ...]
例1:   python 仕訳入力.py あかしまい_04.jpg
例2:   python 仕訳入力.py あかしまい_04.jpg ネオジール_04.pdf
例3:   python 仕訳入力.py *.jpg
"""

import sys
import os
import re
import base64
import json
import warnings
warnings.filterwarnings('ignore')

# Windows文字コード対策
os.environ.setdefault('PYTHONIOENCODING', 'utf-8')

def log(msg):
    print(msg, flush=True)
    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), '仕訳入力ログ.txt'), 'a', encoding='utf-8') as f:
        f.write(msg + '\n')

from datetime import datetime
import anthropic
import openpyxl

# ========== 設定 ==========
ACCOUNTING_FILE = "あかしまい_損益確認 R8年度分.xlsx"  # 会計Excelファイル名
SHEET_SUFFIX = "現金"        # シート名のサフィックス（「あかしまい現金_04」の「現金」部分）
KAKERATE = 1.01              # 掛け率
YEAR_PREFIX = "26"           # 管理番号の年プレフィックス（令和8年→2026→"26"）
API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# 店舗名に含まれるキーワードでカテゴリを上書きするルール（税率分割後にも適用）
# キーワードが店名に含まれていれば、そのカテゴリに強制設定される
STORE_CATEGORY_OVERRIDES = {
    "マクドナルド":           "会議費",
    "マック":                 "会議費",
    "千石電商":               "消耗品費",
    "ポニークリーニング":      "衛生費",
    "JR東日本":               "交通費",
    "JR":                     "交通費",
    "ホットペッパービューティー": "広告宣伝費",
    "ホットペッパー":          "広告宣伝費",
    "リクルート":              "広告宣伝費",
    "NTT東日本":              "通信費",
    "NTT":                    "通信費",
    "So-net":                 "通信費",
    "ソニーネットワーク":      "通信費",
}

# 消費税が明記されない店舗の税率（コードで自動計算）
# キーワード → 税率（%）
STORE_TAX_RATE_RULES = {
    "JR東日本": 10,
    "JR":       10,
    "タクシー": 10,
    "NTT東日本":         10,
    "NTT":               10,
    "So-net":            10,
    "ソニーネットワーク": 10,
}

# メモにこのキーワードが含まれる場合も消費税を計算（店名に関係なく）
MEMO_TAX_RATE_RULES = {
    "Suica":   10,
    "suica":   10,
    "チャージ": 10,
    "スイカ":  10,
}
# ==========================

PROMPT_TEXT = """この画像/ファイルに含まれるレシートをすべて読み取ってください。
レシートが1枚でも複数枚でも、必ず以下のJSON配列形式で返してください。
他のテキストは不要です。

【数字の読み取りについて重要な注意事項】
- 合計金額は「合計」「お買上合計」「ご請求額」などの最終税込金額を読み取ること
- 合計がスタンプ等で読めない場合は 8%対象金額 + 10%対象金額 で計算すること
- 数字は1桁ずつ丁寧に確認し、誤読に注意すること（6と0、1と7、3と8など）
- カタカナも1文字ずつ正確に読むこと（ポとホ、ソとン、シとツ、ウとワなど混同しやすい文字に注意）
- 金額のゼロの個数を必ず数えること（「金１００００円」=10,000円、「金１０００円」=1,000円など桁の読み間違いに注意）
- 全角数字（１２３４）も半角数字と同様に正確に読み取ること
- 消費税は「内消費税」「消費税額」「税額」など括弧書きも含めて必ず読み取ること
  例：「（内消費税 ¥36）」→ 消費税=36、「(内消費税額 ¥148)」→ 消費税=148
- 消費税の記載が全く見当たらない場合のみ0とすること（括弧書きを見落とさないこと）
- タクシー・ハイヤーの領収書は消費税が明記されないが必ず10%課税なので、
  消費税 = int(合計金額 × 10 / 110) で計算して記入すること
  例：合計1,100円 → 消費税=100円、合計550円 → 消費税=50円
- SuicaチャージなどICカードチャージの領収書（JR東日本・駅発行）も交通費として10%で計算すること
  店名はレシートに記載の発行駅・会社名（例：「JR東日本」「日野駅」）を使用すること
  消費税 = int(チャージ金額 × 10 / 110) で計算すること
  チャージ金額は「金１０００円」「金２０００円」「金５０００円」「金１００００円」が一般的。
  ゼロの個数を必ず数えて正確に読み取ること（「金１００００円」=10,000円）
- 「標準税率対象 ¥X」「軽減税率対象 ¥Y」のように税率区分の金額だけ記載され消費税額がない場合は以下で計算すること
  10%のみの場合: 消費税 = int(合計金額 × 10 / 110)
  8%のみの場合:  消費税 = int(合計金額 × 8 / 108)
  例：標準税率対象940円・軽減税率対象0円 → 消費税 = int(940 × 10 / 110) = 85円

【税率内訳フィールドの記入方法】
レシートに8%・10%それぞれの金額が記載されている場合、「税率内訳」に数値を記入してください。
記載がない税率は 税込金額=0、消費税=0 としてください。

よく見られる表記と読み方：
- ファミリーマート: 全角数字で税率が書かれる。以下の順で記載される
    １０％対象日用品等  ¥Y  → 10%の税込金額  ＝ 税率内訳["10%"]["税込金額"]
    （内消費税等        ¥B） → 10%の消費税    ＝ 税率内訳["10%"]["消費税"]
    ８％対象持帰食品等  ¥X  → 8%の税込金額   ＝ 税率内訳["8%"]["税込金額"]
    （内消費税等        ¥A） → 8%の消費税     ＝ 税率内訳["8%"]["消費税"]
  ※「１０％」「８％」は全角数字。カテゴリ名（日用品等・持帰食品等）は無視してよい
  ※（内消費税等）は括弧書きだが必ず読み取ること
  ※読み取り後に必ず検算すること: X + Y = 合計金額 になるはず。一致しない場合は読み直すこと
    例: ８％対象¥4,963 + １０％対象¥632 = ¥5,595（合計と一致）
  ※スタンプ・汚れで片方の税率の金額が読めない場合は逆算すること
    例: 合計¥2,801、８％対象¥2,794 → １０％対象 = 2,801 - 2,794 = ¥7
    逆算した金額でも必ず税率内訳に記入し、省略しないこと

- セブン-イレブン:「税率8%対象商品 ¥X」「税率10%対象商品 ¥Y」→ そのまま税込金額
- ローソン: レシートに以下の順で記載される。各行の意味を正確に対応させること
    (内消費税等 ¥Z)  → 合計税額。絶対に無視すること（10%の金額ではない）
    (10%対象   ¥Y)  → 10%の税込金額  ＝ 税率内訳["10%"]["税込金額"]
    (内消費税額 ¥B)  → 10%の消費税    ＝ 税率内訳["10%"]["消費税"]
    (8%対象    ¥X)  → 8%の税込金額   ＝ 税率内訳["8%"]["税込金額"]
    (内消費税額 ¥A)  → 8%の消費税     ＝ 税率内訳["8%"]["消費税"]
- ツルハドラッグ:「うち、消費税8%対象額 ¥X 税額 ¥A」「うち、消費税10%対象額 ¥Y 税額 ¥B」
- イトーヨーカ堂:「内税率10%対象額 ¥Y（内消費税等 ¥B）」「内税率8%対象額 ¥X（内消費税等 ¥A）」
  ※「内税率X%対象額」の金額は税込金額。「内消費税等」がその消費税額
  旧形式:「(8%)対象計 ¥X」「消費税 ¥A」「(10%)対象計 ¥Y」「消費税 ¥B」（税抜の場合は消費税を加算して税込に）
- イオン・ピーコックストア（外税形式）: 税抜金額と消費税が別行で記載される
    外税 8%対象額  ¥X  → 8%の【税抜】金額（そのまま税込金額に使わないこと）
    外税 8%        ¥A  → 8%の消費税額
    外税10%対象額  ¥Y  → 10%の【税抜】金額（そのまま税込金額に使わないこと）
    外税10%        ¥B  → 10%の消費税額
  ※必ず以下の計算で税込金額を求めること（税抜金額をそのまま税込金額にしないこと）
    税率内訳["8%"]["税込金額"]  = X + A （例: 2,914 + 233 = 3,147）
    税率内訳["8%"]["消費税"]   = A
    税率内訳["10%"]["税込金額"] = Y + B （例: 4 + 0 = 4）
    税率内訳["10%"]["消費税"]  = B

[
  {
    "日付": "YYYY/MM/DD形式（不明な場合は今日の日付）",
    "店名": "店舗名（「株式会社」「有限会社」などの法人格は除き、一般的に呼ばれる店舗名のみ記載。支店名・店舗名がある場合は含めてよい。必ずレシートに印刷された文字を1文字ずつ正確に読み取ること。他の似た店名や記憶・推測で補完せず、レシートの文字そのままを使うこと）",
    "合計金額": 数値のみ（税込合計）,
    "消費税": 数値のみ（合計消費税、不明な場合は0）,
    "カテゴリ": "食材費/交通費/消耗品費/会議費/水道光熱費/通信費/衛生費/広告宣伝費/預り金/雑費/その他 から最も適切なもの（郵便局・ゆうちょ・郵便・宅配便は通信費。特別区民税・都民税・住民税・固定資産税・森林環境税など税金の納付領収書は預り金。合計金額が1,000円未満でどのカテゴリにも当てはまらない場合は雑費。レストラン・食堂・定食屋・居酒屋・カフェ・ファストフード・弁当屋など外食・飲食店での支払いは会議費。クリーニング店・洗濯・衛生用品の購入は衛生費）",
    "メモ": "特記事項があれば（なければ空文字）",
    "税率内訳": {
      "8%": {"税込金額": 数値（8%対象の税込合計、なければ0）, "消費税": 数値（8%分の税額、なければ0）},
      "10%": {"税込金額": 数値（10%対象の税込合計、なければ0）, "消費税": 数値（10%分の税額、なければ0）}
    }
  }
]

レシートが1枚の場合も必ず配列（[ ]）で返してください。"""


def apply_store_category(receipt):
    """店舗名に応じてカテゴリ・消費税を上書きする"""
    store = receipt.get("店名", "")
    result = dict(receipt)

    # カテゴリ上書き
    for keyword, category in STORE_CATEGORY_OVERRIDES.items():
        if keyword in store:
            result["カテゴリ"] = category
            break

    # 消費税が0の場合にコードで計算（店名またはメモのキーワードで判定）
    tax = int(result.get("消費税", 0) or 0)
    total = int(result.get("合計金額", 0) or 0)
    memo = result.get("メモ", "") or ""
    if tax == 0 and total > 0:
        rate = None
        for keyword, r in STORE_TAX_RATE_RULES.items():
            if keyword in store:
                rate = r
                break
        if rate is None:
            for keyword, r in MEMO_TAX_RATE_RULES.items():
                if keyword in memo:
                    rate = r
                    break
        if rate is not None:
            result["消費税"] = int(total * rate / (100 + rate))

    return result


def split_by_tax_rate(receipt):
    """税率内訳をもとに8%・10%を別エントリに分割して返す（Pythonで処理）"""
    breakdown = receipt.get("税率内訳", {})
    r8  = breakdown.get("8%",  {})
    r10 = breakdown.get("10%", {})
    amt8  = int(r8.get("税込金額",  0) or 0)
    amt10 = int(r10.get("税込金額", 0) or 0)
    tax8  = int(r8.get("消費税",   0) or 0)
    tax10 = int(r10.get("消費税",  0) or 0)

    # 検算: 8% + 10% の合計が receipt の合計金額と一致するか確認
    total = int(receipt.get("合計金額", 0) or 0)
    if total > 0 and (amt8 + amt10) > 0 and abs((amt8 + amt10) - total) > 2:
        log(f"  ⚠ 税率内訳の検算不一致: 8%({amt8}円) + 10%({amt10}円) = {amt8+amt10}円 ≠ 合計{total}円（要確認）")

    if amt8 > 0 and amt10 > 0:
        base = {k: v for k, v in receipt.items() if k != "税率内訳"}
        original_cat = receipt.get("カテゴリ", "")
        if original_cat == "会議費":
            # 外食店は分割後も両方会議費
            cat8, cat10 = "会議費", "会議費"
        else:
            # 8%は食材費、10%は消耗品費（10円未満は雑費）
            cat8  = "食材費"
            cat10 = "雑費" if amt10 < 10 else "消耗品費"
        entry8  = {**base, "合計金額": amt8,  "消費税": tax8,  "カテゴリ": cat8,  "メモ": "軽減税率8%"}
        entry10 = {**base, "合計金額": amt10, "消費税": tax10, "カテゴリ": cat10, "メモ": "標準税率10%"}
        return [entry8, entry10]
    else:
        return [receipt]


def parse_filename(filepath):
    """ファイル名から会社名と月を抽出する
    対応形式: 会社名_MM.pdf / 会社名_MM_2.pdf / 会社名_MM_追加.pdf など
    """
    basename = os.path.splitext(os.path.basename(filepath))[0]
    match = re.match(r'^(.+)_(\d{2})(_.*)?$', basename)
    if not match:
        raise ValueError(
            f"ファイル名の形式が正しくありません: {filepath}\n"
            f"正しい形式: 会社名_月.pdf（例: あかしまい_04.pdf、あかしまい_04_2.pdf）"
        )
    company = match.group(1)
    month = match.group(2)
    return company, month


def date_str_to_excel(date_str):
    """YYYY/MM/DD形式の文字列をExcelシリアル値に変換"""
    try:
        d = datetime.strptime(date_str, "%Y/%m/%d")
        excel_epoch = datetime(1899, 12, 30)
        return (d - excel_epoch).days
    except Exception:
        return None


def read_receipts_from_file(filepath):
    """Claude APIでレシートファイルを読み取り、レシートのリストを返す"""
    client = anthropic.Anthropic(api_key=API_KEY)

    with open(filepath, "rb") as f:
        file_data = base64.standard_b64encode(f.read()).decode("utf-8")

    ext = os.path.splitext(filepath)[1].lower()
    log(f"  ファイルを読み取り中: {os.path.basename(filepath)}")

    if ext == ".pdf":
        file_block = {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": file_data},
        }
    else:
        media_types = {
            ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
            ".png": "image/png", ".webp": "image/webp", ".gif": "image/gif",
        }
        file_block = {
            "type": "image",
            "source": {"type": "base64", "media_type": media_types.get(ext, "image/jpeg"), "data": file_data},
        }

    def extract_receipts_from_text(text):
        """テキストからレシートのリストを抽出するヘルパー"""
        # コードブロックを除去
        if "```" in text:
            parts = text.split("```")
            for part in parts:
                if part.startswith("json"):
                    text = part[4:].strip(); break
                elif part.strip().startswith("[") or part.strip().startswith("{"):
                    text = part.strip(); break
        # JSON配列を抽出
        start = text.find("[")
        end = text.rfind("]")
        if start != -1 and end != -1:
            text = text[start:end+1]
        try:
            result = json.loads(text)
        except json.JSONDecodeError as e:
            log(f"  ⚠ JSONパースエラー（自動修復を試みます）: {e}")
            result = []
            decoder = json.JSONDecoder()
            pos = text.find("{")
            while pos != -1 and pos < len(text):
                try:
                    obj, pos_next = decoder.raw_decode(text, pos)
                    result.append(obj)
                    pos = text.find("{", pos_next)
                except json.JSONDecodeError:
                    pos = text.find("{", pos + 1)
            if result:
                log(f"  自動修復成功: {len(result)}件を取得")
        return [result] if isinstance(result, dict) else result

    # 1回目の送信
    messages = [{"role": "user", "content": [file_block, {"type": "text", "text": PROMPT_TEXT}]}]
    data = []
    round_num = 1

    while True:
        response = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=8192,
            messages=messages,
        )
        text = response.content[0].text.strip()
        batch = extract_receipts_from_text(text)
        data.extend(batch)
        log(f"  第{round_num}回取得: {len(batch)}件（累計 {len(data)}件）")

        if response.stop_reason != "max_tokens":
            break  # 正常終了

        # まだ続きがある場合は追加取得
        round_num += 1
        log(f"  ⚠ 途中で切れました。続きを取得します（第{round_num}回）...")
        messages.append({"role": "assistant", "content": text})
        messages.append({"role": "user", "content":
            "まだ読み取っていないレシートがあれば、同じJSON配列形式で出力してください。"
            "すべて読み取り済みであれば [] とだけ返してください。"})

    if isinstance(data, dict):
        data = [data]
    return data


def get_sheet_last_state(ws):
    """シートの最終行の残高と管理番号を取得"""
    last_balance = 0
    last_number = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            last_balance = row[9] if row[9] is not None else 0
            last_number = row[0]
    return last_balance, last_number


def next_kanri_number(last_number, month):
    """次の管理番号を生成（例: 2604001）"""
    prefix = f"{YEAR_PREFIX}{month}"
    if last_number and str(last_number).startswith(prefix):
        seq = int(str(last_number)[-3:]) + 1
    else:
        seq = 1
    return int(f"{prefix}{seq:03d}")


def append_to_sheet(ws, receipts, company):
    """レシートデータをシートに追記"""
    last_balance, last_number = get_sheet_last_state(ws)
    month = str(last_number)[2:4] if last_number else "01"

    added = 0
    for receipt in receipts:
        kanri = next_kanri_number(last_number, month)
        date_val = date_str_to_excel(receipt.get("日付", ""))
        total = int(receipt.get("合計金額", 0))
        tax = int(receipt.get("消費税", 0))
        subtotal = total - tax
        new_balance = last_balance - total
        billing = round(subtotal * KAKERATE, 1)

        ws.append([
            kanri,                        # 管理番号
            None,                         # 店舗（空欄）
            company,                      # 会社
            date_val,                     # 日付
            receipt.get("カテゴリ", ""),   # 内容
            receipt.get("店名", ""),       # 取引先
            total,                        # 税込価格
            subtotal,                     # 税抜価格
            None,                         # 預入（空欄）
            new_balance,                  # 残高
            KAKERATE,                     # 掛け率
            billing,                      # 請求金額
            receipt.get("メモ", ""),       # 備考
        ])

        # 日付列のフォーマットを設定
        last_row = ws.max_row
        date_cell = ws.cell(row=last_row, column=4)
        date_cell.number_format = "YYYY/MM/DD"

        log(f"    [{kanri}] {receipt.get('日付','')} {receipt.get('店名','')} {total:,}円 ({receipt.get('カテゴリ','')})")

        last_balance = new_balance
        last_number = kanri
        added += 1

    return added


def get_last_state_readonly(accounting_path, sheet_name):
    """読み取り専用で最終残高と管理番号を取得（高速）"""
    try:
        wb = openpyxl.load_workbook(accounting_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return 0, None
        ws = wb[sheet_name]
        last_balance = 0
        last_number = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                last_balance = row[9] if row[9] is not None else 0
                last_number = row[0]
        wb.close()
        return last_balance, last_number
    except Exception:
        return 0, None


def create_output_excel(rows, output_path):
    """読み取り結果を別ファイルに出力する"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "仕訳データ"

    headers = ["管理番号", "店舗", "会社", "日付", "内容", "取引先",
               "税込価格", "税抜価格", "預入", "残高", "掛け率", "請求金額", "備考"]
    ws.append(headers)

    from openpyxl.styles import Font, PatternFill, Alignment
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 通貨フォーマットを適用する列（1始まり）: 税込価格=7, 税抜価格=8, 預入=9, 残高=10, 請求金額=12
    currency_cols = [7, 8, 9, 10, 12]
    currency_fmt = '¥#,##0'

    for row in rows:
        ws.append(row)
        r = ws.max_row
        # 日付セルのフォーマット
        ws.cell(row=r, column=4).number_format = "YYYY/MM/DD"
        # 通貨フォーマット
        for col in currency_cols:
            ws.cell(row=r, column=col).number_format = currency_fmt

    # 列幅調整
    widths = [12, 12, 14, 14, 14, 20, 12, 12, 10, 14, 10, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        log("使い方: python 仕訳入力.py 会社名_月.pdf [ファイル2 ...]")
        log("例:     python 仕訳入力.py あかしまい_04.pdf")
        sys.exit(1)

    if not API_KEY:
        log("エラー: ANTHROPIC_API_KEY が設定されていません。")
        sys.exit(1)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    accounting_path = os.path.join(script_dir, ACCOUNTING_FILE)

    if not os.path.exists(accounting_path):
        log(f"エラー: 会計ファイルが見つかりません: {accounting_path}")
        sys.exit(1)

    filepaths = sys.argv[1:]
    all_rows = []
    total_added = 0

    for filepath in filepaths:
        log("=" * 40)
        if not os.path.exists(filepath):
            log(f"スキップ: ファイルが見つかりません: {filepath}")
            continue

        try:
            company, month = parse_filename(filepath)
        except ValueError as e:
            log(f"スキップ: {e}")
            continue

        sheet_name = f"{company}{SHEET_SUFFIX}_{month}"
        log(f"会社: {company}  月: {month}月  → シート: {sheet_name}")
        log(f"最終行を確認中...")

        last_balance, last_number = get_last_state_readonly(accounting_path, sheet_name)
        log(f"  最後の管理番号: {last_number}  残高: {last_balance:,}円")

        try:
            receipts = read_receipts_from_file(filepath)
        except Exception as e:
            log(f"エラー: {filepath} の読み取りに失敗しました: {e}")
            continue

        # 税率内訳をもとに8%・10%を分割し、店舗カテゴリを適用
        expanded = []
        for r in receipts:
            for entry in split_by_tax_rate(r):
                expanded.append(apply_store_category(entry))
        receipts = expanded

        # 重複を自動検出して除外
        seen = set()
        deduped = []
        for r in receipts:
            key = (r.get("日付", ""), r.get("店名", ""), int(r.get("合計金額", 0)))
            if key in seen:
                log(f"  ⚠ 重複を自動除外: {r.get('日付','')} {r.get('店名','')} {int(r.get('合計金額',0)):,}円")
            else:
                seen.add(key)
                deduped.append(r)

        if len(deduped) < len(receipts):
            log(f"  → {len(receipts)}件中 {len(receipts)-len(deduped)}件の重複を除外しました（{len(deduped)}件を登録）")
        receipts = deduped

        # 一覧表示
        log(f"")
        log(f"  ┌─ 登録予定 {len(receipts)}件 " + "─" * 28)
        for i, r in enumerate(receipts, 1):
            log(f"  │ {i:2}. {r.get('日付','')}  {r.get('店名','')[:16]:<16}  {int(r.get('合計金額',0)):>8,}円")
        log(f"  └" + "─" * 38)
        log(f"")

        prev_receipt = None
        current_kanri = None
        for i, receipt in enumerate(receipts, 1):
            # 前のレシートと日付・店名が同じ場合は税率分割とみなし管理番号を引き継ぐ
            is_split = (prev_receipt is not None
                        and receipt.get("日付") == prev_receipt.get("日付")
                        and receipt.get("店名") == prev_receipt.get("店名"))

            if is_split:
                kanri = current_kanri
            else:
                kanri = next_kanri_number(last_number, month)
                current_kanri = kanri
                last_number = kanri

            date_val = date_str_to_excel(receipt.get("日付", ""))
            total = int(receipt.get("合計金額", 0))
            tax = int(receipt.get("消費税", 0))
            subtotal = total - tax
            new_balance = last_balance - total
            billing = round(subtotal * KAKERATE, 1)

            row = [kanri, None, company, date_val, receipt.get("カテゴリ", ""),
                   receipt.get("店名", ""), total, subtotal, None,
                   new_balance, KAKERATE, billing, receipt.get("メモ", "")]
            all_rows.append(row)

            if is_split:
                log(f"  → [{kanri}] {receipt.get('店名','')} {total:,}円 登録（税率分割・同一管理番号）")
            else:
                log(f"  → [{kanri}] {receipt.get('店名','')} {total:,}円 登録")

            last_balance = new_balance
            prev_receipt = receipt
            total_added += 1

    log("=" * 40)

    if total_added == 0:
        log("処理するデータがありませんでした。")
        sys.exit(0)

    # 出力ファイル名（例: 仕訳データ_あかしまい_04.xlsx）
    basename = os.path.splitext(os.path.basename(filepaths[0]))[0]
    output_filename = f"仕訳データ_{basename}.xlsx"
    output_path = os.path.join(script_dir, output_filename)
    create_output_excel(all_rows, output_path)

    log(f"完了！ {total_added}件を {output_filename} に出力しました")
    log(f"")
    log(f"【次の手順】")
    log(f"1. {output_filename} を開く")
    log(f"2. データ行をすべて選択してコピー（Ctrl+C）")
    log(f"3. あかしまい_損益確認 R8年度分.xlsx の {sheet_name} シートを開く")
    log(f"4. 最終行の次のセルに貼り付け（Ctrl+V）")


if __name__ == "__main__":
    main()
